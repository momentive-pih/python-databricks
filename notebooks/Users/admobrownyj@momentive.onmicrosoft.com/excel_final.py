# Databricks notebook source
import pandas as pd
import re
import numpy as np
import glob
import json
import os

# COMMAND ----------

def SQL_connection():
    import pyodbc
    import configparser
    import traceback

    config = configparser.ConfigParser()
      #This configuration path should be configured in Blob storage
    config.read("/dbfs/mnt/momentive-configuration/config-file.ini")
    
    server = config.get('sql_db', 'server')
    database = 'CLD-IT-DEV-PIH-DB1'
    username = config.get('sql_db', 'username')
    password = config.get('sql_db', 'password')
  

    driver= "{ODBC Driver 17 for SQL Server}"
    connection_string = 'DRIVER=' + driver + \
              ';SERVER=' + server + \
              ';PORT=1433' + \
              ';DATABASE=' + database + \
              ';UID=' + username + \
              ';PWD=' + password
    
    try:
      sql_conn = pyodbc.connect(connection_string)
      return sql_conn
      # execute query and save data in pandas df
    except Exception as error:
          print("    \u2717 error message: {}".format(error))
      # I found that traceback prints much more detailed error message
          traceback.print_exc()

# COMMAND ----------

sql_cursor = SQL_connection() 
data = 'select * from momentive.unstructure_processed_data_testing'
data_excel_external_source = pd.read_sql(data, sql_cursor) 

# COMMAND ----------

path = '/dbfs/mnt/momentive-sources-pih/sharepoint-pih/tox-study-pih/'
data_excel_external_source.to_csv(path + 'Unstructured_data.csv', index=False, header=True)

# COMMAND ----------

def reading_excel_sources(source_type, sql_cursor):
  try:
    excel_momentive_source = 'select * from momentive.external_excel_source'
    data_excel_external_source = pd.read_sql(excel_momentive_source, sql_cursor)
    dataframe_excel_sources = data_excel_external_source[(data_excel_external_source['source_type']==source_type) & \
                                                         (data_excel_external_source['is_active_folder']==1) & \
                                                         (data_excel_external_source['is_active_column']=='1') & \
                                                         (data_excel_external_source['is_active_sheet']=='1')]
    
    primary_field = data_excel_external_source[(data_excel_external_source['source_type']==source_type) & \
                                               (data_excel_external_source['is_active_folder']==1) & \
                                               (data_excel_external_source['is_active_column']=='1') & \
                                               (data_excel_external_source['is_active_sheet']=='1') & \
                                               (data_excel_external_source['is_primary']=='1')]
    
    primary_col = primary_field.column_name.values
    print(primary_col)
    external_sheet = list(pd.unique(dataframe_excel_sources['sheet_name']))
    return dataframe_excel_sources, external_sheet, primary_col
  except Exception as e:
    print('Error in reading sources', e)

# COMMAND ----------

def data_validation_to_relevant_non_relevant_split(data_delta, valid_path, primary_column, comp):
  try:
    sql_cursor = SQL_connection()
    query_product_list = 'select * from [momentive].[product_inscope]'
    product_inscope = pd.read_sql(query_product_list, sql_cursor)
    CAS_list = list(product_inscope.cas_no)
    
    nam_prod_list = list(product_inscope.nam_prod)
    bdt_list = list(product_inscope.bdt)
    
    regex1 = re.compile(r'(\d+-\d+-\d+)', re.I) #CAS number formatting
    regex2 = re.compile(r'(Y-\d+)', re.I)  #Y-Number formatting
    regex3 = re.compile(r'tamil', re.I) #
    
    reg_ex = [] 
    reg_ex1 = []
    dbutils.fs.rm((valid_path +'relevant_data_files/').replace("/dbfs",""),True)
    data_delta1 = pd.DataFrame()
    if not data_delta.shape[0]==0:
      for i in range(1,data_delta.shape[0]):
        print('kamal',i)
        product = data_delta.loc[i, primary_column]
        print('product',product)
        reg_ex = regex1.findall(str(product)) or regex2.findall(str(product)) or regex3.findall(str(product))
        reg_ex1 = regex1.findall(str(product)) and regex2.findall(str(product))
        if len(reg_ex1)>0:
          data_copy = data_delta.loc[i,:]
          data_copy.loc[i, primary_column]=reg_ex1[1]
          data_delta =  data_delta.append(data_copy, ignore_index=True)
          data_delta.loc[i, primary_column]=reg_ex1[0]
        elif len(reg_ex)>0:
         if '/' in reg_ex[0]:
              for reg_len in reg_ex[0].split('/') :
                print('reg_len',reg_len)
                data_copy = data_delta.loc[i,:]
                data_copy.loc[i, primary_column]=reg_len
                data_delta =  data_delta.append(data_copy, ignore_index=True)
                #data_delta.loc[i, primary_column]=reg_len 
         else:  
              data_delta.loc[i, primary_column]=reg_ex[0]                
      master_relevant = data_delta.copy()
      master_relevant.rename(columns = {primary_column:'Product'}, inplace=True)
      cas_df = data_delta[primary_column].isin(CAS_list)
      cas_final = data_delta[cas_df]
      cas_final['Product_category'] = 'CAS'
      nam_prod_list_df = data_delta[primary_column].isin(nam_prod_list)
      nam_prod_final = data_delta[nam_prod_list_df]
      nam_prod_final['Product_category'] = 'NAM_PROD'
      bdt_df = data_delta[primary_column].isin(bdt_list)
      bdt_final = data_delta[bdt_df]
      bdt_final['Product_category'] = 'BDT'
      consol_data = pd.concat([cas_final, nam_prod_final, bdt_final])
      print('consol_data.shape',consol_data.shape[0])
      consol_data.rename(columns = {primary_column:'Product'}, inplace=True)
      consol_data['Component'] = comp
      consol_data['is_relevant'] = 1
      master_consol_data = consol_data.copy()
      master_consol_data.drop(columns={'Product_category', 'Component', 'is_relevant'}, inplace=True)
      dbutils.fs.mkdirs((valid_path +'relevant_data_files/').replace("/dbfs","dbfs:")) 
      if not consol_data.shape[0]==0:
        relevant_files = consol_data.to_csv(valid_path + 'relevant_data_files/' + 'relevant_data' +'.csv', index=None, header=True)
      final = master_relevant.append(master_consol_data)
      final.drop_duplicates(keep=False, inplace=True)
      final.reset_index(drop=True, inplace=True)
      final['Component'] = comp
      final['is_relevant'] = 0
      final['Product_category'] = np.nan
      if not final.shape[0]==0:
        final.to_csv(valid_path + 'relevant_data_files/' + 'non_relevant_data' +'.csv', index=None, header=True)  
     # break
  except Exception as e:
    print('Error in exception', e)

# COMMAND ----------

def excel_full_delta_load(valid_path, relevant_data):
#  print('relevant_data', relevant_data.shape[0])
  global data_delta
  data_delta = pd.DataFrame()
  try:
    if not os.path.exists(valid_path + 'valid/'):
      print('kamaal')
      dbutils.fs.mkdirs((valid_path +'valid/').replace("/dbfs","dbfs:")) 
      relevant_data.to_csv(valid_path + 'valid/' + 'valid_data.csv', index=None, header=True, encoding='iso-8859-1')
      match_data = pd.DataFrame()
      flag=1
    else:
      print('murali')
      mat = glob.glob(valid_path + 'valid/' + '*.csv')
      for m in mat:
        match_data = pd.read_csv(m, encoding='iso-8859-1')
        flag=0
#        match_data.to_csv(valid_path + 'valid/' + 'valid_data1.csv')
    match_data.replace({r'[^\x00-\x7F]+':''}, regex=True, inplace=True)
    print('relevant_data count', relevant_data.shape[0])
    print('match_data_count', match_data.shape[0])
    data_delta = relevant_data.append(match_data)
    data_delta.drop_duplicates(keep=False, inplace=True)
    data_delta.reset_index(drop=True, inplace=True)
    print('data_delta', data_delta.shape[0])
    data_delta1 = data_delta.append(match_data)
    dup = data_delta1.duplicated(keep='first')
    data_delta2 = data_delta1[dup]
    data_delta2.reset_index(drop=True, inplace=True)
#    data_delta2.reset_index(drop=True, inplace=True)
    print('data_delta2', data_delta2.shape[0])

    if not data_delta.shape[0]==0 and not flag==1:
      data_to_valid = pd.read_csv(valid_path + 'valid/' + 'valid_data.csv', encoding='iso-8859-1')
      data_to_m = data_delta.append(data_delta2)
      data_to_m.drop_duplicates(keep=False, inplace=True)
      data_to_m.reset_index(drop=True, inplace=True)
      data_to_v = data_to_valid.append(data_to_m)
      data_to_v.reset_index(drop=True, inplace=True)
      dbutils.fs.rm((valid_path +'valid/').replace("/dbfs",""),True)
      data_to_v.to_csv(valid_path + 'valid/' + 'valid_data.csv', index=None, encoding='iso-8859-1')
      data_delta = data_to_m.copy()
#      print('data delta after deletion', data_delta.shape[0])
   
    if not data_delta2.shape[0]==0 and not flag==1:
      data_to_v = data_to_valid.append(data_delta2)
      data_to_v.drop_duplicates(keep=False, inplace=True)
      data_to_v.reset_index(drop=True, inplace=True)
      dbutils.fs.rm((valid_path +'valid/').replace("/dbfs",""),True)
      data_to_v.to_csv(valid_path + 'valid/' + 'valid_data.csv', index=None, encoding='iso-8859-1')
    return data_delta, valid_path
  except Exception as e:
    print('Error in loading', e)

# COMMAND ----------

def reading_excel_data_from_source(valid_path, files, component_data, primary_column, comp):
  global relevant_data
  
  try:
    component_columns = list(component_data['column_name'])
    data_valid_extract = pd.read_csv(files, encoding='iso-8859-1', header=None)
    data_valid_extract = data_valid_extract.dropna(how='all',axis=0)
    data_valid_extract.reset_index(drop=True, inplace=True)
    for i in range(data_valid_extract.shape[0]):
      row_list = list(data_valid_extract.loc[i,:])
      start_row_count = list(set(row_list) & set(component_columns))
      if len(start_row_count) >3:
         value_of_column = i
    valid_data = data_valid_extract[int(value_of_column):]
    valid_data = valid_data.rename(columns=valid_data.iloc[0])
    valid_data.drop(valid_data.index[0], inplace=True)
    valid_data.reset_index(drop=True, inplace=True)
    valid_data.columns = valid_data.columns.str.replace('\n',' ')
    valid_data.columns = valid_data.columns.str.strip()
    valid_data.columns = valid_data.columns.str.replace(r'[^\x00-\x7F]+', '')
    relevant_data = valid_data.loc[:, component_columns]
    relevant_data.replace({r'[^\x00-\x7F]+':''}, regex=True, inplace=True)
    relevant_data.drop_duplicates(keep='first', inplace=True)
    relevant_data.reset_index(drop=True, inplace=True)
    data_delta, valid_path = excel_full_delta_load(valid_path, relevant_data)
    data_validation_to_relevant_non_relevant_split(data_delta, valid_path, primary_column, comp)
  except Exception as e:
    print('Error in reading consolidation', e)

# COMMAND ----------

def unstructure_processed_data(key_value_df_master_data,sql_cursor,cursor):
  query_value = "insert into momentive.unstructure_processed_data (product_type,product,data_extract,is_relevant, created,updated) values ('{}','{}','{}','{}',{},{})"
  print(key_value_df_master_data.shape[0])
  count = 0
  for i in range(key_value_df_master_data.shape[0]):
    try:
      insert_query = query_value.format(key_value_df_master_data['Product_category'][i],\
      key_value_df_master_data['Product'][i],key_value_df_master_data['values'][i].replace("'", "''"),  \
      key_value_df_master_data['is_relevant'][i], 'GETDATE()', 'GETDATE()')
      count = count+1
#      print(count)
      cursor.execute(insert_query)
      sql_cursor.commit()
    except Exception as e:
      print(e)
#      print(insert_query.replace("'", "''"))

# COMMAND ----------

def key_data_extract_external_source(valid_path):
  global key_value_df_master_data
  json_list = []
  try:
    if os.path.exists(valid_path + 'relevant_data_files/'):
      files = glob.glob(valid_path + 'relevant_data_files/' + '*.csv')
      for file in files:
        print(file)
        non_rel_data = pd.read_csv(file, encoding='iso-8859-1')
        product = 'Product'
        temp_data = non_rel_data.copy()
        temp_data.drop([product, 'Product_category', 'Component', 'is_relevant'], axis=1, inplace=True)
        df_dict = temp_data.to_json(orient='records', lines=False, force_ascii=False)
        d = json.loads(df_dict)
        for i in range(len(d)):
          b = json.dumps(d[i], ensure_ascii=False)
          json_list.append(b)
        key_value_df = pd.DataFrame(json_list, columns =['values']) 
        key_value_df_master = non_rel_data.join(key_value_df)
        key_value_df_master_data = key_value_df_master.loc[:, ['Product_category', product, 'values', 'is_relevant']]
        sql_cursor = SQL_connection()
        cursor = sql_cursor.cursor()
        unstructure_processed_data(key_value_df_master_data, sql_cursor, cursor)
  except Exception as e:
    print(e)

# COMMAND ----------

def excel_extract2_key_value_pair(valid_path, sql_cursor):
  try:
    external_source_data = 'select source_type from momentive.external_excel_source'
    source_type_valid = pd.read_sql(external_source_data, sql_cursor)
    list_components = list(pd.unique(source_type_valid.source_type))
    if not list_components:
      print('list is empty')
    else:
      for comp in list_components:
        if comp:
          
          component_data, component_sheet, primary_col = reading_excel_sources(comp.strip(), sql_cursor)
          for sheet in component_sheet:
            print('sheet',sheet)
            if sheet:
              valid_files = glob.glob(valid_path + '*.csv')
              for files in valid_files:
                print('files',files)
                head, tail = os.path.split(files)
                file_name = tail.rsplit('.',1)[0]
                if file_name.strip()==sheet.strip():
                  for primary in primary_col:
                    
                    reading_excel_data_from_source(valid_path, files, component_data, primary, comp)
            else:
              print('sheet is empty')
        #else:
        #  print('component is empty')
  
  except Exception as e:
    print('exception', e)

# COMMAND ----------

path = '/dbfs/mnt/momentive-sources-pih/sharepoint-pih/tox-study-pih/sealants-silanes-library-pih/staging/excel/'

file_list = glob.glob(path +'*.*')
for file in file_list:
  head, tail = os.path.split(file)
  file_name = tail.rsplit('.',1)[0]
  rex = regex1.findall(str(file_name))
  if rex:
     date = str(rex[0])

# COMMAND ----------

import glob
import os
import re
from datetime import datetime
rex1 = re.compile(r'\d{1,2}\s*\/\d{1,2}\s*\/\d{4}') #12/12/2002
rex2 = re.compile(r'[a-zA-Z]*\s*\d{1,2}\s*,\s*\d{4}') #Jan 23, 2002
rex3 = re.compile(r'\d{1,2}\s*\-\s*[a-zA-Z]*\s*\-\s*\d{4}') #12-Jan-2002
rex4 = re.compile(r'\d{1,2}\s*\-\s*[a-zA-Z]*\s*\-\s*\d{2}') #12-Jan-02
rex5 = re.compile(r'\d{1,4}\s*\-\d{1,2}\s*\-\d{2}') #2002-12-12rex5 = re.compile(r'\d{1,4}\s*\-\d{1,2}\s*\-\d{2}') #2002-12-12
rex6 = re.compile(r'\d{1,2}\s*\-\d{1,2}\s*\-\d{4}') #2002-12-12
               
    
#regex1 = re.compile(r'\d{1,2}\s*\/\d{1,2}\s*\/\d{4}","[a-zA-Z]*\s*\d{1,2}\s*,\s*\d{4}","\d{1,2}\s[a-zA-Z-]*\s\d{4}', re.I)
path = '/dbfs/mnt/test-pih/python/'
text_files = glob.glob(path + "*.txt")
from datetime import datetime
flag=0
for text in text_files:
  content = open(text, 'r').read() 
  print(content)
  rex_text = rex4.findall(content) 
  rex_text1 = rex1.findall(content)
  rex_text2 = rex2.findall(content)
  rex_text3 = rex3.findall(content)
  rex_text4 = rex5.findall(content)
  rex_text5 = rex6.findall(content)
  if len(rex_text)>0:
    rex_text.sort(key = lambda date: datetime.strptime(date, '%d-%b-%y').date(), reverse=True) 
    date = rex_text[0]
    print(date)
    flag=1
  elif len(rex_text1)>0:
    rex_text1.sort(key = lambda date: datetime.strptime(date, '%d/%b/%y').date(), reverse=True) 
    date = rex_text1[0]
    print(date)
    flag=1
  elif len(rex_text2)>0:
    rex_text2.sort(key = lambda date: datetime.strptime(date, '%b %d,%Y').date(), reverse=True) 
    date = rex_text2[0]
    print(date)
    flag=1
  elif len(rex_text3)>0:
    rex_text3.sort(key = lambda date: datetime.strptime(date, '%d-%b-%Y').date(), reverse=True) 
    date = rex_text3[0]
    print(date)
    flag=1
  elif len(rex_text4)>0:
    rex_text4.sort(key = lambda date: datetime.strptime(date, '%Y-%m-%d').date(), reverse=True) 
    date = rex_text4[0]
    print(date)
    flag=1
  elif len(rex_text5)>0:
    rex_text5.sort(key = lambda date: datetime.strptime(date, '%d-%m-%Y').date(), reverse=True) 
    date = rex_text5[0]
    print(date)
    flag=1
  else:
    flag=0

if flag==0:
  file_list = glob.glob(path +'*.xlsx')
  print(file_list)
  for file in file_list:
    head, tail = os.path.split(file)
    file_name = tail.rsplit('.',1)[0]
    rex = rex3.findall(str(file_name)) or rex6.findall(str(file_name)) or rex1.findall(str(file_name)) or rex2.findall(str(file_name))\
          or rex4.findall(str(file_name)) or rex5.findall(str(file_name))
    if rex:
      date = str(rex[0])
      print(date)
  
   #    datetime.date()

      

# COMMAND ----------

def excel2csv(path, sheet):
  try:
     # write2csv(path, sheet)
      wb = openpyxl.load_workbook(path)
      sh = wb[sheet]
      head, tail = os.path.split(path)
      filename = path.split('/')[-1].split('.')[0]
      file = head + '/' + 'temp/csv/' + sheet + '.csv'
#    dbutils.fs.rm((absolute_path +'temp/csv/').replace("/dbfs",""),True)
      dbutils.fs.mkdirs((head +'/temp/csv/').replace("/dbfs","dbfs:")) 
      with open(file, 'w', encoding="utf-8") as f:
          c = csv.writer(f)
          for r in sh.rows:
              c.writerow([cell.value for cell in r])
      return file
  except Exception as e:
    print('Error in excel2csv', e)

# COMMAND ----------

import openpyxl
import csv
path = '/dbfs/mnt/momentive-sources-pih/sharepoint-pih/cognizant-pih/sap-bw-pih/staging/excel/Sales_Report.xlsm'
wb = openpyxl.load_workbook(path) 
allsheets = list(wb.sheetnames)
excel2csv(path, allsheets[1])

# COMMAND ----------

def excel_functions():
  valid_path = '/dbfs/mnt/momentive-sources-pih/sharepoint-pih/customer-communications-pih/mpm-2019-pih/analytics/processed/excel/Master Registration List_Cognizant/'
  sql_cursor = SQL_connection()
  excel_extract2_key_value_pair(valid_path, sql_cursor)
  key_data_extract_external_source(valid_path)
#   unstructure_processed_data(query_value,key_value_df_master_data,sql_cursor,cursor)
  #update_operation(query,sql_cursor,cursor)

# COMMAND ----------

excel_functions()

# COMMAND ----------

unstructure_processed_data(key_value_df_master_data,sql_cursor,cursor)
#update_operation(query,sql_cursor,cursor)